import csv
import openpyxl
import json
import sys
from datetime import datetime

def format_date(date_obj):
    if isinstance(date_obj, datetime):
        return date_obj.strftime('%d/%m/%Y')
    return date_obj

def read_csv(file_path, ranges):
    data = []
    with open(file_path, mode='r', encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for index, row in enumerate(csv_reader, start=2):
            if ranges['principal'][0] <= index <= ranges['principal'][1]:
                type_numero = "principal"
            elif ranges['chance'][0] <= index <= ranges['chance'][1]:
                type_numero = "chance"
            else:
                continue
            
            # Utilisation des clés conformes à l'ancien format
            data.append({
                "numero": int(row["Numéros"]),
                "nombreDeSorties": int(row["Nombre de sorties"]),
                "pourcentageDeSorties": row["Pourcentage de sorties"].replace(',', '.'),
                "derniereSortie": row["Dernière sortie"],
                "type": type_numero
            })
    return data

def read_xlsx(file_path, ranges):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {h: i for i, h in enumerate(headers)}
    
    for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if ranges['principal'][0] <= index <= ranges['principal'][1]:
            type_numero = "principal"
        elif ranges['chance'][0] <= index <= ranges['chance'][1]:
            type_numero = "chance"
        else:
            continue

        # Utilisation des clés conformes à l'ancien format
        data.append({
            "numero": row[header_index["Numéros"]].value,
            "nombreDeSorties": row[header_index["Nombre de sorties"]].value,
            "pourcentageDeSorties": str(row[header_index["Pourcentage de sorties"]].value).replace(',', '.'),
            "derniereSortie": format_date(row[header_index["Dernière sortie"]].value),
            "type": type_numero
        })
    return data

if len(sys.argv) < 2:
    print("Usage: python convert.py [jeu]")
    sys.exit(1)

jeu = sys.argv[1].lower()

fichiers = {
    'loto': 'Loto_Stat.csv',
    'euromillions': 'EuroMillions_Stat.csv',
    'eurodreams': 'EuroDreams_Stat.xlsx'
}

ranges = {
    'loto': {'principal': (2, 50), 'chance': (51, 60)},
    'euromillions': {'principal': (2, 51), 'chance': (52, 63)},
    'eurodreams': {'principal': (2, 41), 'chance': (42, 46)}
}

if jeu not in fichiers:
    print(f"Jeu non reconnu: {jeu}")
    sys.exit(2)

file_path = fichiers[jeu]
json_file_path = f"{jeu}_Stat.json"

if file_path.endswith('.csv'):
    game_data = read_csv(file_path, ranges[jeu])
elif file_path.endswith('.xlsx'):
    game_data = read_xlsx(file_path, ranges[jeu])
else:
    print("Format de fichier non pris en charge.")
    sys.exit(3)

with open(json_file_path, mode='w', encoding='utf-8') as json_file:
    json.dump(game_data, json_file, indent=4, ensure_ascii=False)

print(f"Conversion réussie : {json_file_path}")
